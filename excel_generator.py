"""Create GMP gas-test Excel workbooks and matching downloadable chart files."""

import math
from collections import defaultdict
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.data_source import AxDataSource, StrData, StrRef, StrVal
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import MOISTURE_LIMIT, OIL_LIMIT, PARTICLE_LIMITS, TEST_TYPES


def _configure_korean_font():
    """Use an installed Korean font for PNG/PDF exports when one is available."""
    for family in ("Apple SD Gothic Neo", "Malgun Gothic", "NanumGothic", "Noto Sans CJK KR"):
        try:
            font_manager.findfont(family, fallback_to_default=False)
            plt.rcParams["font.family"] = family
            plt.rcParams["axes.unicode_minus"] = False
            return
        except ValueError:
            continue


_configure_korean_font()

RED = PatternFill("solid", fgColor="FF9999")
BLUE = PatternFill("solid", fgColor="D9EAF7")
THIN = Side(style="thin", color="A8B8B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _number(value):
    try:
        return float(str(value).replace(",", "").split()[0])
    except (ValueError, IndexError):
        return 0.0


def _date_key(value):
    parts = [int(part) for part in str(value or "0.0.0").replace("-", ".").split(".") if part.isdigit()]
    return tuple((parts + [0, 0, 0])[:3])


def _sorted(records):
    return sorted(records, key=lambda row: _date_key(row.get("performed_date")), reverse=True)


def _style(cell, fill=None, left=False):
    cell.border = BORDER
    cell.alignment = LEFT if left else CENTER
    if fill:
        cell.fill = fill


def _title(ws, title, end_column, header_row, note=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    ws.cell(1, 1, title).font = Font(bold=True, size=16)
    ws.cell(1, 1).alignment = CENTER
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
        cell = ws.cell(2, 1, note)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[2].height = 72


def _headers(ws, row, headers):
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row, column, header)
        cell.font = Font(bold=True)
        _style(cell, BLUE)


def _widths(ws, widths):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _criteria_number(text, fallback):
    numbers = [float(value.replace(",", "")) for value in __import__("re").findall(r"\d+(?:\.\d+)?", str(text or ""))]
    return numbers[-1] if numbers else fallback


def _matrix(records, value_field, category_fields, selected_grades=None):
    filtered = [row for row in records if not selected_grades or row.get("grade") in selected_grades]
    dates = sorted({str(row.get("performed_date", "")) for row in filtered}, key=_date_key, reverse=True)
    categories = []
    for row in filtered:
        category = tuple(str(row.get(field, "")) for field in category_fields)
        if category not in categories:
            categories.append(category)
    values = {}
    for category in categories:
        for date in dates:
            matches = [
                _number(row.get(value_field)) for row in filtered
                if tuple(str(row.get(field, "")) for field in category_fields) == category
                and str(row.get("performed_date", "")) == date
            ]
            values[(category, date)] = sum(matches) / len(matches) if matches else None
    return categories, dates, values


def _round_up(value, unit):
    return max(unit, math.ceil(value / unit) * unit)


def _chart_major_unit(y_max):
    """Use a small number of clean grid intervals instead of dense black lines."""
    if y_max <= 1:
        return 0.2
    magnitude = 10 ** math.floor(math.log10(y_max))
    return max(magnitude / 10, math.ceil((y_max / 5) / (magnitude / 10)) * (magnitude / 10))


def _set_string_categories(chart, ws, first_row, last_row):
    """Assign cached string labels so Excel renders categories on combo charts."""
    formula = f"'{ws.title}'!$A${first_row}:$A${last_row}"
    points = [StrVal(idx=index, v=str(ws.cell(row, 1).value or "")) for index, row in enumerate(range(first_row, last_row + 1))]
    category = AxDataSource(strRef=StrRef(f=formula, strCache=StrData(ptCount=len(points), pt=points)))
    for chart_part in chart._charts:
        for series in chart_part.series:
            series.cat = category


def _add_excel_chart(ws, chart_row, title, categories, dates, values, limits, y_max, y_axis_title):
    """Write chart source data and add a clustered column chart with limit lines."""
    start = 3
    ws.cell(start, 1, "측정 위치 / 관리번호")
    for offset, date in enumerate(dates, start=2):
        ws.cell(start, offset, date)
    limit_start = 2 + len(dates)
    for offset, (label, _) in enumerate(limits, start=limit_start):
        ws.cell(start, offset, label)

    for row_index, category in enumerate(categories, start=start + 1):
        # A single newline-delimited category is rendered reliably by Excel,
        # unlike a multi-column category range on combined bar/line charts.
        ws.cell(row_index, 1, "\n".join(reversed(category)))
        for offset, date in enumerate(dates, start=2):
            ws.cell(row_index, offset, values[(category, date)])
        for offset, (_, limit) in enumerate(limits, start=limit_start):
            ws.cell(row_index, offset, limit)
        ws.row_dimensions[row_index].height = 30

    data_end = start + len(categories)
    ws.column_dimensions["A"].width = 38
    for column in range(2, limit_start + len(limits)):
        ws.column_dimensions[get_column_letter(column)].width = 16
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = title
    bar.height = 11
    bar.width = min(24, max(16, 10 + len(categories) * 1.6))
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = y_max
    bar.y_axis.majorUnit = _chart_major_unit(y_max)
    bar.y_axis.majorGridlines = None
    bar.x_axis.tickLblPos = "low"
    bar.legend.position = "r"
    if dates:
        bar.add_data(Reference(ws, min_col=2, max_col=1 + len(dates), min_row=start, max_row=data_end), titles_from_data=True)

    if limits:
        line = LineChart()
        line.height = bar.height
        line.width = bar.width
        line.add_data(Reference(ws, min_col=limit_start, max_col=limit_start + len(limits) - 1, min_row=start, max_row=data_end), titles_from_data=True)
        colors = ["E67E22", "C00000", "C07000", "375623", "7030A0"]
        for index, series in enumerate(line.series):
            series.graphicalProperties.line.solidFill = colors[index % len(colors)]
            series.graphicalProperties.line.w = 19050
            series.graphicalProperties.line.prstDash = "dash"
        bar += line
    # openpyxl can reset dimensions while combining charts; set them last.
    bar.height = 11
    bar.width = min(24, max(16, 10 + len(categories) * 1.6))
    _set_string_categories(bar, ws, start + 1, data_end)
    ws.add_chart(bar, chart_row)
    for column in range(1, limit_start + len(limits)):
        ws.cell(start, column).font = Font(bold=True)
        _style(ws.cell(start, column), BLUE)


def _plot_chart(path, title, categories, dates, values, limits, y_max):
    labels = ["\n".join(category) for category in categories]
    figure, axis = plt.subplots(figsize=(16, 8))
    width = 0.75 / max(len(dates), 1)
    positions = list(range(len(categories)))
    for index, date in enumerate(dates):
        data = [values[(category, date)] or 0 for category in categories]
        axis.bar([position - 0.375 + width / 2 + index * width for position in positions], data, width, label=date)
    for index, (label, limit) in enumerate(limits):
        axis.axhline(limit, color=["#e67e22", "#c00000", "#c07000", "#375623", "#7030a0"][index % 5], linestyle="--", label=label)
    axis.set_title(title, fontweight="bold")
    axis.set_ylim(0, y_max)
    axis.set_xticks(positions, labels, fontsize=8)
    axis.legend(loc="center left", bbox_to_anchor=(1, 0.5))
    axis.grid(axis="y", alpha=.25)
    figure.tight_layout()
    figure.savefig(path, dpi=180)
    return figure


def _generate_oil_or_moisture(test_type, records, output_path, chart_paths):
    is_oil = test_type == "oil"
    title = "유분 측정 일지" if is_oil else "수분 측정 일지"
    chart_title = "오일 함량 측정 기록 피벗 차트" if is_oil else "수분 측정 일지 피벗 차트"
    fallback = OIL_LIMIT if is_oil else MOISTURE_LIMIT
    wb = Workbook()
    data = wb.active
    data.title = "데이터"
    _title(data, title, 8, 4)
    headers = ["No.", "관리번호", "측정 위치", "점검결과 (mg/m³)", "측정사진 첨부", "판정", "허용기준", "Performed Date"]
    _headers(data, 4 if is_oil else 3, headers)
    header_row = 4 if is_oil else 3
    for row_index, record in enumerate(_sorted(records), start=header_row + 1):
        values = [record.get("no"), record.get("management_number"), record.get("location"), record.get("result_text"), record.get("photo_attached"), record.get("judgement"), record.get("criteria_text"), record.get("performed_date")]
        for column, value in enumerate(values, start=1):
            cell = data.cell(row_index, column, value)
            _style(cell, left=column in (3, 4, 7))
        if _number(record.get("result_text")) > fallback:
            data.cell(row_index, 4).fill = RED
        if record.get("judgement") != "적합":
            data.cell(row_index, 6).fill = RED
    _widths(data, {"A": 8, "B": 14, "C": 28, "D": 22, "E": 16, "F": 12, "G": 42, "H": 16})
    data.freeze_panes = f"A{header_row + 1}"

    simple = wb.create_sheet("간소화된 데이터")
    _title(simple, title, 6, header_row)
    _headers(simple, header_row, ["No.", "관리번호", "측정 위치", "점검결과 (mg/m³)", "허용기준", "Performed Date"])
    for row_index, record in enumerate(_sorted(records), start=header_row + 1):
        values = [record.get("no"), record.get("management_number"), record.get("location"), _number(record.get("result_text")), _criteria_number(record.get("criteria_text"), fallback), record.get("performed_date")]
        for column, value in enumerate(values, start=1):
            _style(simple.cell(row_index, column, value), left=column == 3)
    _widths(simple, {"A": 8, "B": 14, "C": 28, "D": 22, "E": 16, "F": 16})
    simple.freeze_panes = f"A{header_row + 1}"

    chart_sheet = wb.create_sheet("피벗 차트")
    categories, dates, values = _matrix(_sorted(records), "result_text", ("management_number", "location"))
    limits = [(f"허용기준 ≤ {_criteria_number(record.get('criteria_text'), fallback):g}", _criteria_number(record.get("criteria_text"), fallback)) for record in _sorted(records)]
    limits = list(dict.fromkeys(limits))
    y_max = _round_up(max([value or 0 for value in values.values()] + [value for _, value in limits]), 1 if is_oil else 10)
    _add_excel_chart(
        chart_sheet,
        "G3" if is_oil else "F3",
        chart_title,
        categories,
        dates,
        values,
        limits,
        y_max,
        "점검결과 (mg/m³)",
    )
    figure = _plot_chart(chart_paths[0], chart_title, categories, dates, values, limits, y_max)
    return wb, [figure]


def _generate_airborne(records, output_path, chart_paths):
    wb = Workbook()
    data = wb.active
    data.title = "데이터"
    note = "허용기준 :\n- Grade A: 0.5μm 이상 입자수: 23개/m³ 이하, 5μm 이상 입자수: 4개/m³ 이하\n- Grade B: 0.5μm 이상 입자수: 627개/m³ 이하, 5μm 이상 입자수: 13개/m³ 이하\n- Grade C: 0.5μm 이상 입자수: 23,402개/m³ 이하, 5μm 이상 입자수: 1,540개/m³ 이하\n- Grade D: 0.5μm 이상 입자수: 141,390개/m³ 이하, 5μm 이상 입자수: 8,183개/m³ 이하"
    _title(data, "부유입자 측정 일지", 16, 5, note)
    headers = ["No.", "관리번호", "측정 위치", "Grade", "0.5 μm 이상 부유입자 수/m³", "5.0 μm 이상 부유입자 수/m³", "판정", "Performed Date"]
    headers += [f"{grade} Grade 경고기준 (0.5㎛)" for grade in "ABCD"]
    headers += [f"{grade} Grade 경고기준 (5.0㎛)" for grade in "ABCD"]
    _headers(data, 5, headers)
    for row_index, record in enumerate(_sorted(records), start=6):
        values = [record.get("no"), record.get("management_number"), record.get("location"), record.get("grade"), _number(record.get("particle_05")), _number(record.get("particle_50")), record.get("judgement"), record.get("performed_date")]
        values += [PARTICLE_LIMITS["0.5"][grade] for grade in "ABCD"]
        values += [PARTICLE_LIMITS["5.0"][grade] for grade in "ABCD"]
        for column, value in enumerate(values, start=1):
            _style(data.cell(row_index, column, value), left=column == 3)
        grade = str(record.get("grade", "")).upper()
        if _number(record.get("particle_05")) > PARTICLE_LIMITS["0.5"].get(grade, math.inf):
            data.cell(row_index, 5).fill = RED
        if _number(record.get("particle_50")) > PARTICLE_LIMITS["5.0"].get(grade, math.inf):
            data.cell(row_index, 6).fill = RED
        if record.get("judgement") != "적합":
            data.cell(row_index, 7).fill = RED
    _widths(data, {"A": 8, "B": 14, "C": 28, "D": 9, "E": 24, "F": 24, "G": 12, "H": 16, "I": 20, "J": 20, "K": 20, "L": 20, "M": 20, "N": 20, "O": 20, "P": 20})
    data.freeze_panes = "A6"

    latest_date = max((_date_key(record.get("performed_date")) for record in records), default=(0, 0, 0))
    selected_grades = {"A", "B"} if latest_date[1] == 2 else None
    figures = []
    for particle_size, field, sheet_name, chart_title, image_path in [
        ("0.5", "particle_05", "Pivot 0.5", "PivotChart 0.5 ㎛", chart_paths[0]),
        ("5.0", "particle_50", "Pivot 5.0", "PivotChart 5.0 ㎛", chart_paths[1]),
    ]:
        sheet = wb.create_sheet(sheet_name)
        categories, dates, values = _matrix(_sorted(records), field, ("grade", "management_number", "location"), selected_grades)
        grades = list(dict.fromkeys(category[0] for category in categories))
        limits = [(f"{grade} Grade 경고기준 = {PARTICLE_LIMITS[particle_size][grade]:,}", PARTICLE_LIMITS[particle_size][grade]) for grade in grades if grade in PARTICLE_LIMITS[particle_size]]
        y_max = _round_up(max([value or 0 for value in values.values()] + [value for _, value in limits]), 10)
        _add_excel_chart(
            sheet,
            "A25",
            chart_title,
            categories,
            dates,
            values,
            limits,
            y_max,
            f"{particle_size} μm 이상 부유입자 수/m³",
        )
        figures.append(_plot_chart(image_path, chart_title, categories, dates, values, limits, y_max))
    return wb, figures


def generate_workbook(test_type, records, output_dir, job_id):
    """Generate Excel, PNG charts, and a single PDF containing all charts."""
    output_dir = Path(output_dir)
    chart_paths = [output_dir / f"{job_id}_{test_type}_chart.png"]
    if test_type == "airborne":
        chart_paths = [
            output_dir / f"{job_id}_airborne_pivot_05.png",
            output_dir / f"{job_id}_airborne_pivot_50.png",
        ]
    excel_path = output_dir / f"{job_id}_{TEST_TYPES[test_type]['filename']}"
    if test_type in ("oil", "moisture"):
        workbook, figures = _generate_oil_or_moisture(test_type, records, excel_path, chart_paths)
    else:
        workbook, figures = _generate_airborne(records, excel_path, chart_paths)
    workbook.save(excel_path)
    pdf_path = output_dir / f"{job_id}_{test_type}_charts.pdf"
    with PdfPages(pdf_path) as pdf:
        for figure in figures:
            pdf.savefig(figure)
            plt.close(figure)
    return {"excel": excel_path, "charts": chart_paths, "pdf": pdf_path}
